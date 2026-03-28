"""
SumoSim Tournament Results Exporter

Exports tournament simulation results to CSV and XLSX formats.
Produces two sheets/files:
  1. Full 15-day bout grid (Day, East, West, Winner, Kimarite, East Win%)
  2. Final standings (Rank, Wrestler, Record, KK/MK, Yusho)
"""

from __future__ import annotations

import csv
from pathlib import Path
from typing import Sequence

from data.models import TournamentResult, WrestlerProfile


def _build_bout_rows(
    result: TournamentResult,
    wrestler_map: dict[str, WrestlerProfile],
) -> list[dict]:
    """Build flat bout rows from tournament result."""
    rows = []
    for day in sorted(result.day_results.keys()):
        for bout in result.day_results[day]:
            east = wrestler_map.get(bout.east_id)
            west = wrestler_map.get(bout.west_id)
            winner = wrestler_map.get(bout.winner_id)
            rows.append({
                "Day": day,
                "East": east.shikona if east else bout.east_id,
                "East Rank": east.full_rank if east else "",
                "West": west.shikona if west else bout.west_id,
                "West Rank": west.full_rank if west else "",
                "Winner": winner.shikona if winner else bout.winner_id,
                "Kimarite": bout.predicted_kimarite or "",
                "East Win%": round(bout.east_win_probability * 100, 1),
                "Playoff": "Y" if bout.is_playoff else "",
            })
    return rows


def _build_standings_rows(
    result: TournamentResult,
    wrestler_map: dict[str, WrestlerProfile],
) -> list[dict]:
    """Build standings rows from tournament result."""
    standings = sorted(
        result.final_standings,
        key=lambda s: (-s.wins, s.losses),
    )
    rows = []
    for i, st in enumerate(standings, 1):
        w = wrestler_map.get(st.wrestler_id)
        rows.append({
            "#": i,
            "Wrestler": st.shikona,
            "Rank": w.full_rank if w else "",
            "Heya": w.heya if w else "",
            "Wins": st.wins,
            "Losses": st.losses,
            "Record": st.record,
            "KK/MK": "KK" if st.wins >= 8 else "MK" if st.losses >= 8 else "",
            "Yusho": "🏆" if result.yusho_winner_id == st.wrestler_id else "",
        })
    return rows


def export_csv(
    result: TournamentResult,
    roster: Sequence[WrestlerProfile],
    output_path: str | Path,
) -> Path:
    """Export tournament results to CSV."""
    output_path = Path(output_path)
    wrestler_map = {w.wrestler_id: w for w in roster}

    # Bouts file
    bouts_path = output_path.with_suffix(".csv")
    bout_rows = _build_bout_rows(result, wrestler_map)
    if bout_rows:
        with open(bouts_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=bout_rows[0].keys())
            writer.writeheader()
            writer.writerows(bout_rows)

    # Standings file
    standings_path = output_path.parent / f"{output_path.stem}_standings.csv"
    standings_rows = _build_standings_rows(result, wrestler_map)
    if standings_rows:
        with open(standings_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=standings_rows[0].keys())
            writer.writeheader()
            writer.writerows(standings_rows)

    return bouts_path


def export_xlsx(
    result: TournamentResult,
    roster: Sequence[WrestlerProfile],
    output_path: str | Path,
) -> Path:
    """Export tournament results to a formatted XLSX workbook."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    output_path = Path(output_path).with_suffix(".xlsx")
    wrestler_map = {w.wrestler_id: w for w in roster}
    wb = Workbook()

    # ── Styles ─────────────────────────────────────────────────
    header_font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="8B0000")
    header_align = Alignment(horizontal="center", vertical="center")
    body_font = Font(name="Arial", size=10)
    center = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="C0C0C0"),
        right=Side(style="thin", color="C0C0C0"),
        top=Side(style="thin", color="C0C0C0"),
        bottom=Side(style="thin", color="C0C0C0"),
    )
    win_font = Font(name="Arial", size=10, bold=True, color="006400")
    loss_font = Font(name="Arial", size=10, color="8B0000")
    yusho_fill = PatternFill("solid", fgColor="FFF8DC")
    day_fill = PatternFill("solid", fgColor="F5F0E8")
    kk_fill = PatternFill("solid", fgColor="E8F5E9")
    mk_fill = PatternFill("solid", fgColor="FFEBEE")

    # ── Sheet 1: Bout Grid ─────────────────────────────────────
    ws_bouts = wb.active
    ws_bouts.title = "Bout Grid"

    bout_headers = ["Day", "East", "East Rank", "West", "West Rank",
                     "Winner", "Kimarite", "East Win%"]
    for col, h in enumerate(bout_headers, 1):
        cell = ws_bouts.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    bout_rows = _build_bout_rows(result, wrestler_map)
    prev_day = None
    for r, row_data in enumerate(bout_rows, 2):
        values = [row_data[h] for h in bout_headers]
        current_day = values[0]

        for c, val in enumerate(values, 1):
            cell = ws_bouts.cell(row=r, column=c, value=val)
            cell.font = body_font
            cell.border = thin_border
            if c in (1, 7, 8):
                cell.alignment = center

        # Day separator shading
        if current_day != prev_day and current_day is not None:
            for c in range(1, len(bout_headers) + 1):
                ws_bouts.cell(row=r, column=c).fill = day_fill
        prev_day = current_day

        # Bold the winner
        winner_cell = ws_bouts.cell(row=r, column=6)
        winner_cell.font = Font(name="Arial", size=10, bold=True)

    col_widths = [6, 16, 20, 16, 20, 16, 16, 10]
    for i, w in enumerate(col_widths, 1):
        ws_bouts.column_dimensions[get_column_letter(i)].width = w

    ws_bouts.auto_filter.ref = f"A1:H{len(bout_rows) + 1}"
    ws_bouts.freeze_panes = "A2"

    # ── Sheet 2: Standings ─────────────────────────────────────
    ws_stand = wb.create_sheet("Standings")

    stand_headers = ["#", "Wrestler", "Rank", "Heya", "Wins", "Losses",
                      "Record", "KK/MK", "Yusho"]
    for col, h in enumerate(stand_headers, 1):
        cell = ws_stand.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    standings_rows = _build_standings_rows(result, wrestler_map)
    for r, row_data in enumerate(standings_rows, 2):
        values = [row_data[h] for h in stand_headers]
        for c, val in enumerate(values, 1):
            cell = ws_stand.cell(row=r, column=c, value=val)
            cell.font = body_font
            cell.border = thin_border
            if c in (1, 5, 6, 7, 8, 9):
                cell.alignment = center

        # Color kachi-koshi / make-koshi
        kk_mk = row_data["KK/MK"]
        if kk_mk == "KK":
            ws_stand.cell(row=r, column=7).font = win_font
            ws_stand.cell(row=r, column=8).font = win_font
            for c in range(1, len(stand_headers) + 1):
                ws_stand.cell(row=r, column=c).fill = kk_fill
        elif kk_mk == "MK":
            ws_stand.cell(row=r, column=7).font = loss_font
            ws_stand.cell(row=r, column=8).font = loss_font
            for c in range(1, len(stand_headers) + 1):
                ws_stand.cell(row=r, column=c).fill = mk_fill

        # Highlight yusho winner
        if row_data["Yusho"]:
            for c in range(1, len(stand_headers) + 1):
                ws_stand.cell(row=r, column=c).fill = yusho_fill
            ws_stand.cell(row=r, column=2).font = Font(
                name="Arial", size=10, bold=True, color="B8860B"
            )

    stand_widths = [5, 16, 20, 14, 6, 7, 8, 7, 7]
    for i, w in enumerate(stand_widths, 1):
        ws_stand.column_dimensions[get_column_letter(i)].width = w

    ws_stand.freeze_panes = "A2"

    # ── Sheet 3: Summary ───────────────────────────────────────
    ws_summary = wb.create_sheet("Summary")
    ws_summary["A1"] = "SumoSim Tournament Projection"
    ws_summary["A1"].font = Font(name="Arial", size=14, bold=True)
    ws_summary["A3"] = "Basho"
    ws_summary["B3"] = result.basho_id
    ws_summary["A4"] = "Yusho Winner"
    yusho_w = wrestler_map.get(result.yusho_winner_id)
    ws_summary["B4"] = yusho_w.shikona if yusho_w else result.yusho_winner_id or "TBD"
    ws_summary["A5"] = "Total Bouts"
    ws_summary["B5"] = sum(len(b) for b in result.day_results.values())
    ws_summary["A6"] = "Days Simulated"
    ws_summary["B6"] = len(result.day_results)

    if result.playoff_results:
        ws_summary["A8"] = "Playoff Bouts"
        ws_summary["A8"].font = Font(name="Arial", bold=True)
        for i, bout in enumerate(result.playoff_results):
            east = wrestler_map.get(bout.east_id)
            west = wrestler_map.get(bout.west_id)
            winner = wrestler_map.get(bout.winner_id)
            ws_summary[f"A{9+i}"] = f"{east.shikona if east else bout.east_id} vs {west.shikona if west else bout.west_id}"
            ws_summary[f"B{9+i}"] = f"{winner.shikona if winner else bout.winner_id}"

    for col in ["A", "B"]:
        ws_summary.column_dimensions[col].width = 28

    ws_summary["A3"].font = Font(name="Arial", bold=True)
    ws_summary["A4"].font = Font(name="Arial", bold=True)
    ws_summary["A5"].font = Font(name="Arial", bold=True)
    ws_summary["A6"].font = Font(name="Arial", bold=True)

    wb.save(str(output_path))
    return output_path
